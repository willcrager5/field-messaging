#!/usr/bin/env python3
"""
Generate Bill's launch brief from the Field brand-launch critical-path spreadsheet.

Weekly Monday lookahead: overdue + this week + next week preview.

Output:
  Markdown file in <out>/YYYY-MM-DD-weekly.md, voiced as Field Brand Strategist.
"""
import argparse
import datetime as dt
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.stderr.write("openpyxl required. pip install openpyxl --break-system-packages\n")
    sys.exit(2)

LAUNCH_DATE = dt.date(2026, 5, 5)


def parse_date(val):
    """Accept datetime, date, or stringy dates like 'Apr 16' or '2026-04-16'."""
    if val is None:
        return None
    if isinstance(val, dt.datetime):
        return val.date()
    if isinstance(val, dt.date):
        return val
    s = str(val).strip()
    if not s:
        return None
    # Try a few formats. Assume current year if only month/day.
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%b %d, %Y", "%B %d, %Y"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    for fmt in ("%b %d", "%B %d"):
        try:
            d = dt.datetime.strptime(s, fmt).date()
            return d.replace(year=LAUNCH_DATE.year)
        except ValueError:
            continue
    return None


def extract_bill_tasks(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    tasks = []
    # Critical Path is a summary rollup — its items duplicate detail-tab tasks.
    # Skip it to avoid double-counting.
    SKIP = {"Critical Path"}
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP:
            continue
        ws = wb[sheet_name]
        # Find header row — row whose col A is "#" or "Phase"
        header_row = None
        for r in range(1, min(ws.max_row + 1, 8)):
            v = ws.cell(r, 1).value
            if v in ("#", "Phase"):
                header_row = r
                break
        if not header_row:
            continue
        headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
        # Map columns
        def col(*names):
            for n in names:
                if n in headers:
                    return headers.index(n) + 1
            return None

        task_col = col("Key Milestone", "Task")
        due_col = col("Due Date")
        status_col = col("Status")
        owner_col = col("Owner")
        notes_col = col("Notes", "Notes / Open Questions")
        if not all([task_col, due_col, status_col, owner_col]):
            continue

        for r in range(header_row + 1, ws.max_row + 1):
            owner = ws.cell(r, owner_col).value
            if not owner or "bill" not in str(owner).lower():
                continue
            status = (ws.cell(r, status_col).value or "").strip()
            if status.lower() in ("complete", "completed", "done"):
                continue
            tasks.append({
                "sheet": sheet_name,
                "task": ws.cell(r, task_col).value,
                "due": parse_date(ws.cell(r, due_col).value),
                "due_raw": ws.cell(r, due_col).value,
                "status": status or "Not Started",
                "owner": owner,
                "notes": ws.cell(r, notes_col).value if notes_col else None,
            })
    return tasks


def bucket(tasks, today):
    """Split tasks into overdue / today / this_week / next_week / later / undated."""
    end_of_week = today + dt.timedelta(days=(6 - today.weekday()))  # Sunday
    end_of_next_week = end_of_week + dt.timedelta(days=7)
    buckets = {
        "overdue": [],
        "today": [],
        "this_week": [],
        "next_week": [],
        "later": [],
        "undated": [],
    }
    for t in tasks:
        d = t["due"]
        if d is None:
            buckets["undated"].append(t)
        elif d < today:
            buckets["overdue"].append(t)
        elif d == today:
            buckets["today"].append(t)
        elif d <= end_of_week:
            buckets["this_week"].append(t)
        elif d <= end_of_next_week:
            buckets["next_week"].append(t)
        else:
            buckets["later"].append(t)
    for k in buckets:
        buckets[k].sort(key=lambda x: (x["due"] or dt.date.max, x["sheet"]))
    return buckets


def fmt_task(t, show_sheet=True):
    due = t["due"].strftime("%b %-d") if t["due"] else "no date"
    name = (t["task"] or "").strip()
    line = f"- **{due}** — {name}"
    if show_sheet:
        line += f"  _({t['sheet']})_"
    # Short note reminder if task is vague
    if t["notes"] and len(name) < 60:
        note = str(t["notes"]).strip()
        if note and len(note) < 120:
            line += f"  \n  {note}"
    return line


def days_to_launch(today):
    return (LAUNCH_DATE - today).days


def pick_push(buckets):
    """Pick the single task Bill should push on. Prefer messaging/positioning
    work across all near-term buckets before falling back to whatever's oldest."""
    pools = buckets["overdue"] + buckets["today"] + buckets["this_week"]
    for t in pools:
        name = (t["task"] or "").lower()
        if any(w in name for w in ("messaging framework", "value prop", "pillar", "elevator", "how we talk")):
            return t
    for t in pools:
        name = (t["task"] or "").lower()
        if any(w in name for w in ("messaging", "positioning")):
            return t
    return pools[0] if pools else None


def compose_weekly(buckets, today):
    n = days_to_launch(today)
    lines = []
    lines.append(f"# Bill's brief — week of {today.strftime('%b %-d')}\n")
    lines.append(f"{n} days to launch.\n")

    if buckets["overdue"]:
        lines.append("## Slipped\n")
        lines.append("These were due before today. Pick them up or reassign.\n")
        for t in buckets["overdue"]:
            lines.append(fmt_task(t))
        lines.append("")

    if buckets["today"] or buckets["this_week"]:
        lines.append("## This week\n")
        for t in buckets["today"] + buckets["this_week"]:
            lines.append(fmt_task(t))
        lines.append("")
    else:
        lines.append("## This week\n")
        lines.append("Nothing on you this week. Use it.\n")

    if buckets["next_week"]:
        lines.append("## Next week — get ahead\n")
        for t in buckets["next_week"]:
            lines.append(fmt_task(t))
        lines.append("")

    push = pick_push(buckets)
    if push:
        lines.append("## The one to move today\n")
        name = (push["task"] or "").strip()
        lines.append(f"**{name}** _({push['sheet']})._")
        lines.append("Everything downstream waits on this. Start here.\n")

    if buckets["undated"]:
        lines.append("## No date on the sheet\n")
        for t in buckets["undated"]:
            lines.append(fmt_task(t))
        lines.append("")

    return "\n".join(lines).strip() + "\n"


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", required=True)
    p.add_argument("--out", required=True)
    p.add_argument("--today", help="Override today (YYYY-MM-DD)")
    args = p.parse_args()

    today = dt.date.fromisoformat(args.today) if args.today else dt.date.today()
    tasks = extract_bill_tasks(args.xlsx)
    buckets = bucket(tasks, today)
    brief = compose_weekly(buckets, today)

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / f"{today.isoformat()}-weekly.md"
    out_file.write_text(brief)
    print(f"Wrote {out_file}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
